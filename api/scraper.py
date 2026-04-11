import re
import os
import time
from datetime import datetime
from urllib.parse import urlparse, quote
from collections import Counter

import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

# 둘 중 하나만 쓰면 된다.
# 1) webdriver_manager 자동 사용
from webdriver_manager.chrome import ChromeDriverManager

# 2) 로컬 chromedriver 직접 경로 사용하고 싶으면 아래 CHROMEDRIVER_PATH에 넣기
CHROMEDRIVER_PATH = None
# 예:
# CHROMEDRIVER_PATH = "/usr/local/bin/chromedriver"

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference


ALLOWED_HOSTNAME = "m.place.naver.com"
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/120.0.0.0 Safari/537.36"
)

POSITIVE_WORDS = [
    "맛있", "친절", "좋", "깨끗", "추천", "훌륭", "최고", "재방문",
    "만족", "가성비", "신선", "부드럽", "정성", "빠르", "넉넉",
    "쾌적", "세련", "고소", "담백", "깔끔", "푸짐", "든든", "상냥"
]

NEGATIVE_WORDS = [
    "별로", "불친절", "느리", "비싸", "짜다", "싱겁", "더럽",
    "실망", "최악", "재방문 안", "기다림", "불편", "시끄럽",
    "텁텁", "차갑", "오래", "부족", "아쉽", "실수", "불쾌",
    "눅눅", "딱딱", "질기", "비위생", "형편없"
]


class ScrapeError(Exception):
    pass


def is_naver_place_url(url: str) -> bool:
    try:
        parsed = urlparse(url)
    except Exception:
        return False

    if parsed.scheme not in ["http", "https"]:
        return False

    return parsed.netloc.lower() == ALLOWED_HOSTNAME


def fetch_html(url: str, timeout: int = 10, retries: int = 3) -> str:
    if not is_naver_place_url(url):
        raise ScrapeError("Not a m.place.naver.com URL")

    headers = {
        "User-Agent": USER_AGENT,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    }

    last_exc = None
    for attempt in range(1, retries + 1):
        try:
            resp = requests.get(url, headers=headers, timeout=timeout)
            if resp.status_code != 200:
                raise ScrapeError(f"Failed to fetch page, status={resp.status_code}")
            return resp.text
        except Exception as e:
            last_exc = e
            if attempt == retries:
                raise ScrapeError(f"fetch_html failed after {retries} attempts: {e}")

    raise ScrapeError(f"fetch_html failed: {last_exc}")


def _build_driver_service():
    if CHROMEDRIVER_PATH:
        return Service(CHROMEDRIVER_PATH)
    return Service(ChromeDriverManager().install())


def fetch_html_selenium(url: str, timeout: int = 20) -> str:
    if not is_naver_place_url(url):
        raise ScrapeError("Not a m.place.naver.com URL")

    options = Options()
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    # 디버깅 중이면 headless 끄는 걸 추천
    # options.add_argument("--headless")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument(f"user-agent={USER_AGENT}")
    options.binary_location = "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"

    driver = None

    try:
        print("👉 드라이버 생성")
        chrome_service = _build_driver_service()
        driver = webdriver.Chrome(service=chrome_service, options=options)

        print("👉 크롬 실행 성공")
        driver.set_page_load_timeout(timeout)
        driver.get(url)

        print("👉 페이지 로딩 완료")
        time.sleep(3)

        # 더보기 버튼 반복 클릭
        for _ in range(10):
            try:
                more_button = WebDriverWait(driver, 3).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, "a.fvwqf"))
                )
                driver.execute_script("arguments[0].click();", more_button)
                time.sleep(1)
            except Exception:
                break

        print("👉 더보기 클릭 완료")

        # 리뷰 DOM 로드 대기
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located(
                (
                    By.CSS_SELECTOR,
                    "li.place_apply_pui.EjjAW, li.EjjAW, li._3oG8X, "
                    "li.review_item, div._1km0z, div._3QPE6, div._3Rixz"
                )
            )
        )

        time.sleep(1)
        return driver.page_source

    except Exception as e:
        print("🔥 내부 에러:", e)
        raise ScrapeError(f"fetch_html_selenium failed: {e}")

    finally:
        if driver:
            try:
                driver.quit()
            except Exception:
                pass


def _extract_text(node):
    if node is None:
        return ""
    return node.get_text(separator=" ", strip=True)


def _extract_rating(item):
    rating = None

    star = item.select_one("div.review_score div.star_score span")
    if star is not None:
        star_text = star.get("style", "")
        m = re.search(r"width:\s*(\d+)%", star_text)
        if m:
            rating = float(m.group(1)) / 20.0

    if rating is None:
        score_node = item.select_one("span.score, span._1D7MQ")
        if score_node:
            txt = score_node.get_text(strip=True)
            try:
                rating = float(txt)
            except Exception:
                pass

    return rating


def _normalize_review_text(text):
    if not text:
        return text

    text = text.strip()

    m_tail = re.search(r"^(.*?)(?:\s*반응 남기기|\s*방문일).*", text)
    if m_tail:
        text = m_tail.group(1).strip()

    m = re.search(r"^[^\n]+리뷰[^\n]*?\s+(.*)$", text)
    if m and len(m.group(1).strip()) > 5:
        candidate = m.group(1).strip()
        m_tail2 = re.search(r"^(.*?)(?:\s*반응 남기기|\s*방문일).*", candidate)
        if m_tail2:
            candidate = m_tail2.group(1).strip()
        if len(candidate) <= len(text):
            text = candidate

    return text


def _normalize_review_date(date_text):
    if not date_text:
        return date_text

    date_text = date_text.strip()

    m = re.search(r"(\d{4})년\s*(\d{1,2})월\s*(\d{1,2})일\s*([월화수목금토일]+요일)", date_text)
    if m:
        yyyy = int(m.group(1))
        mm = int(m.group(2))
        dd = int(m.group(3))
        dow = m.group(4)
        return f"{yyyy:04d}-{mm:02d}-{dd:02d} {dow}"

    m2 = re.search(r"(\d{2})\.(\d{1,2})\.(\d{1,2})\.([월화수목금토일])", date_text)
    if m2:
        prefix = int(m2.group(1))
        yyyy = 2000 + prefix if prefix < 70 else 1900 + prefix
        mm = int(m2.group(2))
        dd = int(m2.group(3))
        wik = m2.group(4)
        return f"{yyyy:04d}-{mm:02d}-{dd:02d} {wik}요일"

    return date_text


def _extract_tags(item):
    tags = []
    tag_selectors = [
        "div.pui__HLNvmI",
        "span._2-GE-",
        "div.review_tag",
        "div._1I3nE span",
    ]
    for sel in tag_selectors:
        for node in item.select(sel):
            txt = _extract_text(node)
            if txt and txt not in tags:
                tags.append(txt)
    return tags


def analyze_sentiment_and_keywords(text: str):
    pos_words = []
    neg_words = []

    for w in POSITIVE_WORDS:
        if w in text:
            pos_words.append(w)

    for w in NEGATIVE_WORDS:
        if w in text:
            neg_words.append(w)

    if len(pos_words) > len(neg_words):
        sentiment = "positive"
    elif len(neg_words) > len(pos_words):
        sentiment = "negative"
    else:
        sentiment = "neutral"

    return sentiment, pos_words, neg_words


def parse_naver_place_reviews(html: str):
    soup = BeautifulSoup(html, "html.parser")

    candidates = []
    selectors = [
        "li.place_apply_pui.EjjAW",
        "li.EjjAW",
        "li._3oG8X",
        "li.review_item",
        "div.review_item",
        "div._1km0z",
        "div._3QPE6",
        "div._3Rixz",
        "div.section_review_list li",
    ]

    for sel in selectors:
        found = soup.select(sel)
        if found:
            candidates = found
            break

    if not candidates:
        return {"count": 0, "reviews": []}

    reviews = []

    for item in candidates:
        text = ""
        author = ""
        date = ""
        rating = None
        tags = []

        for x in [
            "div.pui__vn15t2",
            "div.review_text",
            "span._1W6Y3",
            "span._3oJ5G",
            "p._1i3d0",
            "div._1U3gJ",
            "div._2tkPb",
        ]:
            node = item.select_one(x)
            if node:
                text = _extract_text(node)
                if text:
                    break

        if not text:
            text = _extract_text(item.select_one("p") or item)

        text = _normalize_review_text(text)

        author_node = item.select_one(
            "span.pui__uslU0d, span._3hl2F, span._1Gy50, div._3-1M1, span._2s0fL"
        )
        if author_node:
            author = _extract_text(author_node)

        date_node = item.select_one(
            "span.pui__gfuUIT, span._3fM31, span._1Q9DG, time, span._2Aa_p"
        )
        if date_node:
            date = _normalize_review_date(_extract_text(date_node))

        rating = _extract_rating(item)
        tags = _extract_tags(item)

        if not text:
            continue

        sentiment, pos_words, neg_words = analyze_sentiment_and_keywords(text)

        reviews.append({
            "author": author,
            "text": text,
            "tags": tags,
            "date": date,
            "rating": rating,
            "sentiment": sentiment,
            "pos_words": pos_words,
            "neg_words": neg_words,
        })

    return {
        "count": len(reviews),
        "reviews": reviews,
    }


def save_reviews_to_excel(reviews, directory="testData"):
    os.makedirs(directory, exist_ok=True)
    filename = datetime.now().strftime("%y%m%d_%H%M%S") + ".xlsx"
    filepath = os.path.join(directory, filename)

    wb = Workbook()

    # 1. 리뷰 원본 시트
    ws = wb.active
    ws.title = "reviews"
    ws.append([
        "date", "author", "rating", "sentiment",
        "pos_words", "neg_words", "tags", "text"
    ])

    for r in reviews:
        ws.append([
            r.get("date", ""),
            r.get("author", ""),
            r.get("rating", ""),
            r.get("sentiment", ""),
            ", ".join(r.get("pos_words", [])),
            ", ".join(r.get("neg_words", [])),
            ", ".join(r.get("tags", [])),
            r.get("text", ""),
        ])

    # 2. 키워드 통계 시트
    ws2 = wb.create_sheet("keyword_summary")

    pos_counter = Counter()
    neg_counter = Counter()

    for r in reviews:
        pos_counter.update(r.get("pos_words", []))
        neg_counter.update(r.get("neg_words", []))

    ws2.append(["positive_keyword", "count", "", "negative_keyword", "count"])

    max_len = max(len(pos_counter), len(neg_counter), 1)
    pos_items = pos_counter.most_common()
    neg_items = neg_counter.most_common()

    for i in range(max_len):
        pos_key, pos_val = ("", "")
        neg_key, neg_val = ("", "")

        if i < len(pos_items):
            pos_key, pos_val = pos_items[i]
        if i < len(neg_items):
            neg_key, neg_val = neg_items[i]

        ws2.append([pos_key, pos_val, "", neg_key, neg_val])

    # 3. 날짜별 감성 추이 시트
    ws3 = wb.create_sheet("trend")

    trend = {}
    for r in reviews:
        d = r.get("date", "")
        if not d:
            d = "unknown"

        date_key = d.split(" ")[0]

        if date_key not in trend:
            trend[date_key] = {"positive": 0, "negative": 0, "neutral": 0}

        trend[date_key][r.get("sentiment", "neutral")] += 1

    ws3.append(["date", "positive", "negative", "neutral"])

    for d in sorted(trend.keys()):
        ws3.append([
            d,
            trend[d]["positive"],
            trend[d]["negative"],
            trend[d]["neutral"],
        ])

    # 4. 추이 그래프
    chart = LineChart()
    chart.title = "Sentiment Trend by Date"
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Date"
    chart.height = 10
    chart.width = 18

    data = Reference(ws3, min_col=2, max_col=4, min_row=1, max_row=ws3.max_row)
    cats = Reference(ws3, min_col=1, min_row=2, max_row=ws3.max_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    ws3.add_chart(chart, "F2")

    # 열 너비 조정
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["F"].width = 25
    ws.column_dimensions["G"].width = 25
    ws.column_dimensions["H"].width = 100

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["D"].width = 20
    ws2.column_dimensions["E"].width = 10

    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["B"].width = 10
    ws3.column_dimensions["C"].width = 10
    ws3.column_dimensions["D"].width = 10

    wb.save(filepath)
    return filepath


def scrape_naver_place_reviews(url: str, use_selenium: bool = False):
    if use_selenium:
        html = fetch_html_selenium(url)
    else:
        html = fetch_html(url)

    result = parse_naver_place_reviews(html)

    reviews = result.get("reviews", [])
    print(f"[INFO] Total reviews found: {len(reviews)}")

    for i, r in enumerate(reviews, start=1):
        print(
            f"[REVIEW {i}] "
            f"author={r.get('author')!r}, "
            f"date={r.get('date')!r}, "
            f"rating={r.get('rating')!r}, "
            f"sentiment={r.get('sentiment')!r}, "
            f"pos_words={r.get('pos_words')!r}, "
            f"neg_words={r.get('neg_words')!r}"
        )
        print(f"          content={r.get('text')!r}")

    excel_path = save_reviews_to_excel(reviews, directory=os.path.join(os.getcwd(), "testData"))
    print(f"[INFO] Reviews saved to Excel: {excel_path}")

    return result

def search_naver_places(query: str, timeout: int = 30) -> list:
    """
    네이버 지도에서 업체/가게를 검색합니다 (Selenium 사용).
    검색 결과 페이지에서 정보를 추출합니다.
    
    Args:
        query: 검색어 (예: '아카시아')
        timeout: 요청 타임아웃 (초)
    
    Returns:
        검색 결과 리스트. 각 항목은:
        {
            'place_id': '12345678',
            'name': '업체명',
            'address': '주소',
            'category': '카테고리',
            'url': 'https://m.place.naver.com/place/{place_id}'
        }
    """
    if not query or not query.strip():
        raise ScrapeError("Search query cannot be empty")
    
    # 검색어 URL 인코딩
    encoded_query = quote(query.strip(), safe='')
    search_url = f"https://m.map.naver.com/search?query={encoded_query}"
    
    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument(f"user-agent={USER_AGENT}")
    
    try:
        chrome_service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=chrome_service, options=options)
        driver.set_page_load_timeout(timeout)
        
        print(f"[INFO] Navigating to {search_url}")
        driver.get(search_url)
        
        # 검색 결과 로드 대기 (최대 15초)
        try:
            WebDriverWait(driver, 15).until(
                EC.presence_of_all_elements_located(
                    (By.CSS_SELECTOR, "li.place_item, div[class*='place'], a[href*='/place/']")
                )
            )
        except Exception:
            # 위 선택자가 없으면 잠시 대기
            time.sleep(10)
        
        html = driver.page_source
        soup = BeautifulSoup(html, "html.parser")
        results = []
        
        # 모바일 앱 버전의 검색 결과 항목들을 찾기
        place_items = soup.select(
            "li.place_item, "
            "li[class*='place'], "
            "div[class*='PlaceItem'], "
            "div[class*='search_item'], "
            "li[class*='search_result']"
        )
        
        print(f"[DEBUG] Found {len(place_items)} place items")
        
        # place_items가 없으면 모든 링크에서 place 관련 검색
        if not place_items:
            all_links = soup.select("a[href*='/place/']")
            print(f"[DEBUG] Found {len(all_links)} links with '/place/'")
            
            seen_ids = set()
            for link in all_links:
                href = link.get('href', '')
                place_id_match = re.search(r'/place/(\d+)', href)
                if place_id_match and place_id_match.group(1) not in seen_ids:
                    place_id = place_id_match.group(1)
                    seen_ids.add(place_id)
                    
                    name = ""
                    address = ""
                    category = ""
                    
                    # link의 부모 <li>에서 추가 정보 추출
                    li_parent = link.find_parent("li")
                    if li_parent:
                        # name 추출: <strong class="_item_name_..."> 태그에서
                        name_tag = li_parent.select_one("strong[class*='_item_name']")
                        if name_tag:
                            name = name_tag.get_text(strip=True)
                        
                        # category 추출: <em class="_item_category_..."> 태그에서
                        category_tag = li_parent.select_one("em[class*='_item_category']")
                        if category_tag:
                            category = category_tag.get_text(strip=True)
                        
                        # address 추출: <button class="_item_address_..."> 태그에서
                        # button의 text는 "<i>주소 보기</i>실제_주소" 형태이므로, i 태그를 제거하고 나머지 가져오기
                        address_btn = li_parent.select_one("button[class*='_item_address']")
                        if address_btn:
                            btn_text = address_btn.get_text(strip=True)
                            # "주소 보기" 제거
                            address = btn_text.replace("주소 보기", "").replace("주소보기", "").strip()
                    
                    results.append({
                        'place_id': place_id,
                        'name': name,
                        'address': address,
                        'category': category,
                        'url': f'https://m.place.naver.com/place/{place_id}'
                    })
        else:
            # place_items가 있으면 각 항목에서 정보 추출
            for item in place_items:
                try:
                    # 링크 추출
                    name_link = item.select_one("a[href*='/place/']")
                    if not name_link:
                        continue
                    
                    href = name_link.get('href', '')
                    place_id_match = re.search(r'/place/(\d+)', href)
                    if not place_id_match:
                        continue
                    
                    place_id = place_id_match.group(1)
                    name = name_link.get_text(strip=True)
                    address = ""
                    category = ""
                    
                    results.append({
                        'place_id': place_id,
                        'name': name,
                        'address': address,
                        'category': category,
                        'url': f'https://m.place.naver.com/place/{place_id}'
                    })
                except Exception as e:
                    print(f"[WARN] Failed to parse search result item: {e}")
                    continue
        
        if not results:
            print(f"[WARN] No search results found for query: {query}")
        
        return results
    except Exception as e:
        raise ScrapeError(f"search_naver_places failed: {e}")
    finally:
        try:
            driver.quit()
        except Exception:
            pass


def _extract_place_details(driver, detail_url: str, timeout: int = 20) -> tuple:
    """
    상세 페이지에서 업체명, 주소, 카테고리를 추출합니다.
    
    Returns:
        (name, address, category) 튜플
    """
    try:
        print(f"[DEBUG] Loading detail page: {detail_url}")
        driver.get(detail_url)
        time.sleep(1)  # 페이지 로드 대기 (더 짧은 대기)
        
        html = driver.page_source
        soup = BeautifulSoup(html, "html.parser")
        
        name = ""
        address = ""
        category = ""
        
        # 업체명 추출
        # 페이지 제목에서 " | 지도 | 네이버" 부분 제거
        title_elem = soup.select_one("h1, h2, [class*='title']")
        if title_elem:
            name = title_elem.get_text(strip=True)
        
        if not name:
            # title 태그에서 추출
            title_tag = soup.find("title")
            if title_tag:
                title_text = title_tag.get_text(strip=True)
                # " | 지도 | 네이버" 제거
                name = title_text.split(" | ")[0].strip()
        
        # 주소와 카테고리 추출 - 텍스트에서 직접 찾기
        all_text = soup.get_text()
        
        # 주소 추출 (정규식으로 "시 구 동" 패턴 찾기)
        address_patterns = re.findall(r'([가-힣]+시(?:\s+[가-힣]+구)?(?:\s+[가-힣]+동)?(?:\s+[^,\n]{0,30})?)', all_text)
        if address_patterns:
            address = address_patterns[0].strip()[:100]  # 최대 100자
        
        # 카테고리 추출 - 일반적인 카테고리 키워드 찾기
        category_keywords = ["카페", "음식점", "식당", "레스토랑", "펍", "바", "병원", "약국", 
                            "미용실", "헬스", "짐", "학원", "도서관", "영화관", "호텔", "펜션",
                            "카지노", "극장", "박물관", "공원", "은행", "편의점", "마트"]
        for keyword in category_keywords:
            if keyword in all_text:
                category = keyword
                break
        
        print(f"[DEBUG] Extracted: name={name[:30] if name else ''}..., address={address[:30] if address else ''}..., category={category}")
        return (name, address, category)
    except Exception as e:
        print(f"[DEBUG] Failed to extract place details: {e}")
        return ("", "", "")

if __name__ == "__main__":
    url = "https://m.place.naver.com/restaurant/1316435683/review/visitor"

    print("=== 크롤링 시작 ===")

    try:
        result = scrape_naver_place_reviews(url, use_selenium=True)

        print("\n=== 결과 요약 ===")
        print("리뷰 개수:", result["count"])

        for i, r in enumerate(result["reviews"], start=1):
            print(f"\n[리뷰 {i}]")
            print("작성자:", r["author"])
            print("날짜:", r["date"])
            print("평점:", r["rating"])
            print("감성:", r["sentiment"])
            print("긍정 키워드:", r["pos_words"])
            print("부정 키워드:", r["neg_words"])
            print("내용:", r["text"])

    except Exception as e:
        print("에러 발생:", e)